from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
import database as db
from utils import generate_code_from_korean, generate_code_from_english
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
import calendar as _cal
from datetime import date as _date

app = Flask(__name__)
app.secret_key = 'payroll-secret-key-2025'

with app.app_context():
    db.init_db()


# ─── 포맷 필터 ───────────────────────────────────────────
def fmt_number(value):
    try:
        return f"{int(value):,}"
    except (ValueError, TypeError):
        return value

app.jinja_env.filters['fmt_number'] = fmt_number


# ─── 엑셀 스타일 헬퍼 ────────────────────────────────────
def _header_style(ws, row, cols, fill_color='1a4f8a'):
    fill = PatternFill('solid', fgColor=fill_color)
    font = Font(bold=True, color='FFFFFF', size=10)
    align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _sample_style(ws, row, cols):
    fill = PatternFill('solid', fgColor='F0F4F8')
    font = Font(color='718096', size=10, italic=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border


# ════════════════════════════════════════════════════════
#  대시보드
# ════════════════════════════════════════════════════════
@app.route('/')
def index():
    employees = db.get_all_employees()
    allowances = db.get_all_allowances()
    active_count = sum(1 for e in employees if not e['resign_date'])
    resign_count = sum(1 for e in employees if e['resign_date'])
    return render_template('index.html',
                           total_emp=len(employees),
                           active_count=active_count,
                           resign_count=resign_count,
                           allowance_count=len(allowances))


# ════════════════════════════════════════════════════════
#  인원 등록 — Step 1: 한글 이름 암호화
# ════════════════════════════════════════════════════════
@app.route('/employees/register')
def employee_register():
    return render_template('employees/register.html')


@app.route('/employees/register/template/step1')
def download_step1_template():
    """Step1 양식 다운로드 (이름 입력용)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '이름목록'
    headers = ['이름(한글)']
    ws.append(headers)
    _header_style(ws, 1, range(1, 2))
    ws.append(['홍길동'])
    ws.append(['홍길동a'])
    ws.append(['남궁민호'])
    _sample_style(ws, 2, [1])
    _sample_style(ws, 3, [1])
    _sample_style(ws, 4, [1])
    ws.column_dimensions['A'].width = 20
    ws['A1'].comment = None
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='01_이름목록_양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/employees/register/step1', methods=['POST'])
def process_step1():
    """한글 이름 엑셀 업로드 → 암호화 → 결과 다운로드"""
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx)을 업로드해주세요.', 'danger')
        return redirect(url_for('employee_register'))

    try:
        wb_in = openpyxl.load_workbook(file)
        ws_in = wb_in.active
        rows = list(ws_in.iter_rows(min_row=2, values_only=True))
        names = [str(r[0]).strip() for r in rows if r[0] and str(r[0]).strip()]

        if not names:
            flash('이름 데이터가 없습니다.', 'warning')
            return redirect(url_for('employee_register'))

        # 암호화
        results = []
        seen_codes = {}
        for name in names:
            code = generate_code_from_korean(name)
            if code in seen_codes:
                flash(f'중복 코드 발생: {name} → {code} (기존: {seen_codes[code]})', 'warning')
            seen_codes[code] = name
            results.append((name, code))

        # 결과 엑셀 생성
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = '암호화결과'
        ws_out.append(['원본이름(한글)', '암호화코드'])
        _header_style(ws_out, 1, [1, 2])
        for orig, code in results:
            ws_out.append([orig, code])
        ws_out.column_dimensions['A'].width = 20
        ws_out.column_dimensions['B'].width = 20

        buf = io.BytesIO()
        wb_out.save(buf)
        buf.seek(0)
        flash(f'{len(results)}명 암호화 완료. 파일을 다운로드하세요.', 'success')
        return send_file(buf, as_attachment=True,
                         download_name='02_암호화결과.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        flash(f'파일 처리 중 오류: {str(e)}', 'danger')
        return redirect(url_for('employee_register'))


# ════════════════════════════════════════════════════════
#  인원 등록 — Step 2: 연봉 등록
# ════════════════════════════════════════════════════════
@app.route('/employees/register/template/step2')
def download_step2_template():
    """Step2 양식 다운로드 (코드+연봉)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '연봉등록'
    ws.append(['암호화코드', '연봉(원)'])
    _header_style(ws, 1, [1, 2])
    ws.append(['h4g3d411', 36000000])
    _sample_style(ws, 2, [1, 2])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='03_연봉등록_양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/employees/register/step2', methods=['POST'])
def process_step2():
    """암호화코드 + 연봉 업로드 → DB 등록"""
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx)을 업로드해주세요.', 'danger')
        return redirect(url_for('employee_register'))

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        ok, skipped, errors = 0, 0, []
        for i, row in enumerate(rows, start=2):
            if not row[0]:
                continue
            code = str(row[0]).strip()
            try:
                salary = int(str(row[1]).replace(',', '').replace(' ', ''))
            except (ValueError, TypeError):
                errors.append(f'{i}행: 연봉 값 오류 ({row[1]})')
                continue

            existing = db.get_employee_by_code(code)
            if existing:
                db.update_employee_salary(code, salary)
                ok += 1
            else:
                # 신규 등록 (기본값으로, 연봉만 입력)
                db.create_employee({
                    'name_code':           code,
                    'department':          '',
                    'contract_type':       '정규직',
                    'join_date':           None,
                    'resign_date':         None,
                    'severance_type':      'DB',
                    'scheduled_hours':     209.0,
                    'overtime_hours':      0.0,
                    'annual_salary':       salary,
                    'base_salary':         0,
                    'ordinary_wage':       0,
                    'match_key':           '',
                    'research_tax_exempt': 'N',
                    'child_deduction':     'N',
                    'child_count':         0,
                    'is_exception':        'N',
                    'employee_type':       '직원',
                })
                ok += 1

        msg = f'{ok}명 연봉 등록/업데이트 완료.'
        if errors:
            msg += f' 오류 {len(errors)}건: ' + ' / '.join(errors[:3])
            flash(msg, 'warning')
        else:
            flash(msg, 'success')

    except Exception as e:
        flash(f'파일 처리 중 오류: {str(e)}', 'danger')

    return redirect(url_for('employee_register'))


# ════════════════════════════════════════════════════════
#  인원 관리 — 목록 / 개별 수정 / 삭제
# ════════════════════════════════════════════════════════
@app.route('/employees')
def employee_list():
    employees = db.get_all_employees()
    return render_template('employees/list.html', employees=employees)


@app.route('/employees/new', methods=['GET', 'POST'])
def employee_new():
    allowances  = db.get_all_allowances()
    deductions  = db.get_all_deductions()
    if request.method == 'POST':
        data = _parse_employee_form(request.form)
        if not data['name_code']:
            flash('이름 코드를 입력해주세요.', 'danger')
            return render_template('employees/form.html', employee=data,
                                   allowances=allowances, deductions=deductions, mode='new')
        emp_id = db.create_employee(data)
        for a in allowances:
            amt_str = request.form.get(f'allowance_{a["id"]}', '0').replace(',', '')
            db.upsert_employee_allowance(emp_id, a['id'], int(amt_str or 0))
        for d in deductions:
            amt_str = request.form.get(f'deduction_{d["id"]}', '0').replace(',', '')
            db.upsert_employee_deduction(emp_id, d['id'], int(amt_str or 0))
        flash(f'인원이 등록되었습니다. [{data["name_code"]}]', 'success')
        return redirect(url_for('employee_list'))
    return render_template('employees/form.html', employee=None,
                           allowances=allowances, deductions=deductions, mode='new')


@app.route('/employees/<int:emp_id>/edit', methods=['GET', 'POST'])
def employee_edit(emp_id):
    employee = db.get_employee(emp_id)
    if not employee:
        flash('존재하지 않는 인원입니다.', 'danger')
        return redirect(url_for('employee_list'))
    allowances    = db.get_all_allowances()
    deductions    = db.get_all_deductions()
    emp_allowances = {ea['allowance_item_id']: ea['amount']
                      for ea in db.get_employee_allowances(emp_id)}
    emp_deductions = {ed['deduction_item_id']: ed['amount']
                      for ed in db.get_employee_deductions(emp_id)}
    if request.method == 'POST':
        data = _parse_employee_form(request.form)
        db.update_employee(emp_id, data)
        for a in allowances:
            amt_str = request.form.get(f'allowance_{a["id"]}', '0').replace(',', '')
            db.upsert_employee_allowance(emp_id, a['id'], int(amt_str or 0))
        for d in deductions:
            amt_str = request.form.get(f'deduction_{d["id"]}', '0').replace(',', '')
            db.upsert_employee_deduction(emp_id, d['id'], int(amt_str or 0))
        _recalc_ordinary_wage(emp_id, allowances, emp_allowances)
        flash('인원 정보가 수정되었습니다.', 'success')
        return redirect(url_for('employee_list'))
    ordinary_wage_calc = _calc_ordinary_wage(employee, allowances, emp_allowances)
    monthly_pay = round(employee['annual_salary'] / 12) if employee['annual_salary'] else 0
    overtime_pay_calc = monthly_pay - (employee['base_salary'] or 0)
    return render_template('employees/form.html', employee=employee,
                           allowances=allowances, emp_allowances=emp_allowances,
                           deductions=deductions, emp_deductions=emp_deductions,
                           ordinary_wage_calc=ordinary_wage_calc,
                           overtime_pay_calc=max(0, overtime_pay_calc),
                           mode='edit')


@app.route('/employees/<int:emp_id>/delete', methods=['POST'])
def employee_delete(emp_id):
    emp = db.get_employee(emp_id)
    if emp:
        db.delete_employee(emp_id)
        flash(f'인원이 삭제되었습니다. [{emp["name_code"]}]', 'success')
    return redirect(url_for('employee_list'))


@app.route('/employees/bulk-delete', methods=['POST'])
def employee_bulk_delete():
    ids = [int(i) for i in request.form.getlist('delete_ids') if i.isdigit()]
    if ids:
        db.bulk_delete_employees(ids)
    flash(f'{len(ids)}명이 삭제되었습니다.', 'success' if ids else 'warning')
    return redirect(url_for('employee_list'))


def _calc_ordinary_wage(employee, allowances, emp_allowances):
    """통상임금 = 기본급 + is_ordinary_wage=1 수당 합계"""
    total = employee['base_salary'] or 0
    for a in allowances:
        if a['is_ordinary_wage'] and a['id'] in emp_allowances:
            total += emp_allowances[a['id']]
    return total


def _recalc_ordinary_wage(emp_id, allowances, emp_allowances):
    """편집 완료 후 통상임금을 DB에 저장"""
    emp = db.get_employee(emp_id)
    if not emp:
        return
    # 최신 수당 다시 읽기
    fresh_ea = {ea['allowance_item_id']: ea['amount']
                for ea in db.get_employee_allowances(emp_id)}
    ow = emp['base_salary'] or 0
    for a in allowances:
        if a['is_ordinary_wage'] and a['id'] in fresh_ea:
            ow += fresh_ea[a['id']]
    import sqlite3 as _s3
    conn = db.get_db()
    conn.execute("UPDATE employees SET ordinary_wage=? WHERE id=?", (ow, emp_id))
    conn.commit()
    conn.close()


def _parse_employee_form(form):
    child_ded = 'Y' if form.get('child_deduction') == 'Y' else 'N'
    annual_salary    = int(form.get('annual_salary', '0').replace(',', '') or 0)
    scheduled_hours  = float(form.get('scheduled_hours', 209) or 209)
    overtime_hours   = float(form.get('overtime_hours', 0) or 0)
    ordinary_wage    = int(form.get('ordinary_wage', '0').replace(',', '') or 0)
    base_salary_raw  = int(form.get('base_salary', '0').replace(',', '') or 0)

    # 기본급 자동계산: (연봉/12) / (소정+연장) × 소정
    total_hours = scheduled_hours + overtime_hours
    if annual_salary > 0 and total_hours > 0:
        base_salary = int((annual_salary / 12) / total_hours * scheduled_hours)
    else:
        base_salary = base_salary_raw

    return {
        'name_code':            form.get('name_code', '').strip(),
        'department':           form.get('department', '').strip(),
        'contract_type':        form.get('contract_type', '정규직'),
        'join_date':            form.get('join_date', '') or None,
        'resign_date':          form.get('resign_date', '') or None,
        'severance_type':       form.get('severance_type', 'DB'),
        'scheduled_hours':      scheduled_hours,
        'overtime_hours':       overtime_hours,
        'annual_salary':        annual_salary,
        'base_salary':          base_salary,
        'ordinary_wage':        ordinary_wage,
        'match_key':            form.get('match_key', '').strip(),
        'research_tax_exempt':  form.get('research_tax_exempt', 'N'),
        'child_deduction':      child_ded,
        'child_count':          int(form.get('child_count', 0) or 0) if child_ded == 'Y' else 0,
        'is_exception':         form.get('is_exception', 'N'),
        'employee_type':        form.get('employee_type', '직원'),
    }


# ════════════════════════════════════════════════════════
#  인원 관리 — 일괄 업로드
# ════════════════════════════════════════════════════════
@app.route('/employees/bulk-upload')
def employee_bulk_upload():
    return render_template('employees/bulk_upload.html')


@app.route('/employees/bulk-upload/template')
def download_bulk_template():
    """인원관리 일괄업로드 템플릿 (기등록 인원 pre-populate)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '인원정보'
    headers = [
        '암호화코드', '부서/팀', '계약유형', '입사일(YYYY-MM-DD)',
        '퇴사일(없으면공백)', '퇴직금유형', '소정근로시간(월)', '연장근로시간(월)',
        '매칭키(주민번호앞7자리)', '연구보조비비과세(Y/N)',
        '자녀공제대상(Y/N)', '자녀수(0-3)', '예외자여부(Y/N)', '직원임원구분'
    ]
    ws.append(headers)
    _header_style(ws, 1, range(1, len(headers) + 1))

    employees = db.get_all_employees()
    thin = Side(style='thin', color='CCCCCC')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if employees:
        for row_idx, emp in enumerate(employees, start=2):
            # 날짜: YYYY-MM-DD 앞 10자리만
            join_d  = (str(emp['join_date'])[:10]  if emp['join_date']  else '')
            resign_d = (str(emp['resign_date'])[:10] if emp['resign_date'] else '')
            row_data = [
                emp['name_code'],
                emp['department']        or '',
                emp['contract_type']     or '정규직',
                join_d,
                resign_d,
                emp['severance_type']    or 'DB',
                emp['scheduled_hours']   if emp['scheduled_hours'] is not None else 209,
                emp['overtime_hours']    if emp['overtime_hours']  is not None else 0,
                emp['match_key']         or '',
                emp['research_tax_exempt'] or 'N',
                emp['child_deduction']   or 'N',
                emp['child_count']       or 0,
                emp['is_exception']      or 'N',
                emp['employee_type']     or '직원',
            ]
            ws.append(row_data)
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = data_border
    else:
        # 등록된 인원 없을 때 예시 행
        sample = [
            'h4g3d411', '영업팀', '정규직', '2020-01-15',
            '', 'DB', '209', '0',
            '8501012', 'N', 'N', '0', 'N', '직원'
        ]
        ws.append(sample)
        _sample_style(ws, 2, range(1, len(headers) + 1))

    col_widths = [18, 14, 12, 20, 16, 12, 14, 14, 22, 20, 16, 12, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='04_인원정보_일괄업로드_양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/employees/bulk-upload/process', methods=['POST'])
def process_bulk_upload():
    import re as _re
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx)을 업로드해주세요.', 'danger')
        return redirect(url_for('employee_bulk_upload'))

    VALID_CONTRACT      = {'정규직', '계약직', '파트타임'}
    VALID_SEVERANCE     = {'DB', 'DC', '1년미만'}
    VALID_EMP_TYPE      = {'직원', '임원'}
    DATE_RE             = _re.compile(r'^\d{4}-\d{2}-\d{2}$')

    def _val(v, default=''):
        return str(v).strip() if v is not None else default

    def _yn(v):
        return 'Y' if str(v).strip().upper() == 'Y' else 'N'

    def _date(v):
        """값이 있으면 앞 10자리만, 없으면 None. 형식 오류 시 (None, 오류메시지) 반환"""
        s = _val(v)
        if not s:
            return None, None
        s = s[:10]
        if not DATE_RE.match(s):
            return None, f'날짜 형식 오류: {v} → YYYY-MM-DD 필요'
        return s, None

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        ok, error_rows = 0, []

        for i, row in enumerate(rows, start=2):
            if not row[0]:
                continue

            code = _val(row[0])
            row_errors = []

            # ── 계약유형
            contract_type = _val(row[2], '정규직')
            if contract_type not in VALID_CONTRACT:
                row_errors.append(f'계약유형 오류 "{contract_type}" → 정규직/계약직/파트타임')

            # ── 퇴직금유형
            severance_type = _val(row[5], 'DB')
            if severance_type not in VALID_SEVERANCE:
                row_errors.append(f'퇴직금유형 오류 "{severance_type}" → DB/DC/1년미만')

            # ── 직원임원구분
            employee_type = _val(row[13], '직원')
            if employee_type not in VALID_EMP_TYPE:
                row_errors.append(f'직원임원구분 오류 "{employee_type}" → 직원/임원')

            # ── 날짜
            join_date,   e1 = _date(row[3])
            resign_date, e2 = _date(row[4])
            if e1: row_errors.append(f'입사일 {e1}')
            if e2: row_errors.append(f'퇴사일 {e2}')

            # ── 근로시간
            try:
                scheduled_hours = float(_val(row[6], '209') or 209)
                if scheduled_hours <= 0:
                    row_errors.append('소정근로시간은 0보다 커야 합니다')
            except ValueError:
                row_errors.append(f'소정근로시간 숫자 오류: {row[6]}')
                scheduled_hours = 209

            try:
                overtime_hours = float(_val(row[7], '0') or 0)
            except ValueError:
                row_errors.append(f'연장근로시간 숫자 오류: {row[7]}')
                overtime_hours = 0

            # ── 자녀공제
            child_ded = _yn(row[10])
            try:
                child_count = int(_val(row[11], '0') or 0)
                if not (0 <= child_count <= 3):
                    row_errors.append(f'자녀수 오류 {child_count} → 0~3')
                    child_count = 0
            except ValueError:
                row_errors.append(f'자녀수 숫자 오류: {row[11]}')
                child_count = 0
            if child_ded == 'Y' and child_count == 0:
                row_errors.append('자녀공제 대상(Y)인데 자녀수가 0')

            # ── 오류 있으면 저장하지 않고 기록
            if row_errors:
                error_rows.append({'row': i, 'code': code, 'errors': row_errors})
                continue

            # ── 기존 인원 salary 보존
            existing = db.get_employee_by_code(code)
            try:
                data = {
                    'name_code':           code,
                    'department':          _val(row[1]),
                    'contract_type':       contract_type,
                    'join_date':           join_date,
                    'resign_date':         resign_date,
                    'severance_type':      severance_type,
                    'scheduled_hours':     scheduled_hours,
                    'overtime_hours':      overtime_hours,
                    'annual_salary':       existing['annual_salary'] if existing else 0,
                    'base_salary':         existing['base_salary']   if existing else 0,
                    'ordinary_wage':       existing['ordinary_wage'] if existing else 0,
                    'match_key':           _val(row[8]),
                    'research_tax_exempt': _yn(row[9]),
                    'child_deduction':     child_ded,
                    'child_count':         child_count if child_ded == 'Y' else 0,
                    'is_exception':        _yn(row[12]),
                    'employee_type':       employee_type,
                }
                db.upsert_employee(data)
                ok += 1
            except Exception as e:
                error_rows.append({'row': i, 'code': code, 'errors': [str(e)]})

        if error_rows:
            flash(f'{ok}명 등록/업데이트 완료. {len(error_rows)}행에서 오류가 발생했습니다.', 'warning')
            return render_template('employees/bulk_upload.html',
                                   error_rows=error_rows, ok_count=ok)

        flash(f'{ok}명 등록/업데이트 완료.', 'success')
        return redirect(url_for('employee_list'))

    except Exception as e:
        flash(f'파일 처리 중 오류: {str(e)}', 'danger')
        return redirect(url_for('employee_bulk_upload'))


# ════════════════════════════════════════════════════════
#  수당 항목 관리
# ════════════════════════════════════════════════════════
@app.route('/allowances')
def allowance_list():
    allowances = db.get_all_allowances()
    return render_template('allowances/list.html', allowances=allowances)


@app.route('/allowances/new', methods=['GET', 'POST'])
def allowance_new():
    if request.method == 'POST':
        data = _parse_allowance_form(request.form)
        if not data['name']:
            flash('수당명을 입력해주세요.', 'danger')
            return render_template('allowances/form.html', item=data, mode='new')
        db.create_allowance(data)
        flash(f'수당 항목이 추가되었습니다. [{data["name"]}]', 'success')
        return redirect(url_for('allowance_list'))
    return render_template('allowances/form.html', item=None, mode='new')


@app.route('/allowances/<int:item_id>/edit', methods=['GET', 'POST'])
def allowance_edit(item_id):
    item = db.get_allowance(item_id)
    if not item:
        flash('존재하지 않는 수당 항목입니다.', 'danger')
        return redirect(url_for('allowance_list'))
    if request.method == 'POST':
        data = _parse_allowance_form(request.form)
        db.update_allowance(item_id, data)
        flash(f'수당 항목이 수정되었습니다. [{data["name"]}]', 'success')
        return redirect(url_for('allowance_list'))
    return render_template('allowances/form.html', item=item, mode='edit')


@app.route('/allowances/<int:item_id>/delete', methods=['POST'])
def allowance_delete(item_id):
    item = db.get_allowance(item_id)
    if item:
        db.delete_allowance(item_id)
        flash(f'수당 항목이 삭제되었습니다. [{item["name"]}]', 'success')
    return redirect(url_for('allowance_list'))


@app.route('/allowances/<int:item_id>/toggle', methods=['POST'])
def allowance_toggle(item_id):
    db.toggle_allowance(item_id)
    return redirect(url_for('allowance_list'))


def _parse_allowance_form(form):
    return {
        'name':              form.get('name', '').strip(),
        'is_ordinary_wage':  1 if form.get('is_ordinary_wage') else 0,
        'severance_db':      1 if form.get('severance_db') else 0,
        'severance_dc':      1 if form.get('severance_dc') else 0,
        'payment_condition': form.get('payment_condition', 'fixed'),
        'condition_value':   int(form.get('condition_value', 0) or 0),
        'apply_target':      form.get('apply_target', 'all'),
        'is_active':         1 if form.get('is_active') else 0,
    }


# ════════════════════════════════════════════════════════
#  공제 항목 관리
# ════════════════════════════════════════════════════════
@app.route('/deductions')
def deduction_list():
    items = db.get_all_deductions()
    return render_template('deductions/list.html', items=items)


@app.route('/deductions/new', methods=['GET', 'POST'])
def deduction_new():
    if request.method == 'POST':
        data = _parse_deduction_form(request.form)
        if not data['name']:
            flash('공제 항목명을 입력해주세요.', 'danger')
            return render_template('deductions/form.html', item=data, mode='new')
        db.create_deduction(data)
        flash(f'공제 항목이 추가되었습니다. [{data["name"]}]', 'success')
        return redirect(url_for('deduction_list'))
    return render_template('deductions/form.html', item=None, mode='new')


@app.route('/deductions/<int:item_id>/edit', methods=['GET', 'POST'])
def deduction_edit(item_id):
    item = db.get_deduction(item_id)
    if not item:
        flash('존재하지 않는 항목입니다.', 'danger')
        return redirect(url_for('deduction_list'))
    if request.method == 'POST':
        data = _parse_deduction_form(request.form)
        db.update_deduction(item_id, data)
        flash(f'공제 항목이 수정되었습니다. [{data["name"]}]', 'success')
        return redirect(url_for('deduction_list'))
    return render_template('deductions/form.html', item=item, mode='edit')


@app.route('/deductions/<int:item_id>/delete', methods=['POST'])
def deduction_delete(item_id):
    item = db.get_deduction(item_id)
    if item:
        db.delete_deduction(item_id)
        flash(f'공제 항목이 삭제되었습니다. [{item["name"]}]', 'success')
    return redirect(url_for('deduction_list'))


@app.route('/deductions/<int:item_id>/toggle', methods=['POST'])
def deduction_toggle(item_id):
    db.toggle_deduction(item_id)
    return redirect(url_for('deduction_list'))


def _parse_deduction_form(form):
    return {
        'name':              form.get('name', '').strip(),
        'is_insurance':      1 if form.get('is_insurance') else 0,
        'payment_condition': form.get('payment_condition', 'fixed'),
        'condition_value':   int(form.get('condition_value', 0) or 0),
        'apply_target':      form.get('apply_target', 'all'),
        'is_active':         1 if form.get('is_active') else 0,
    }


# ════════════════════════════════════════════════════════
#  급여작업
# ════════════════════════════════════════════════════════

def _get_active_employees_for_period(year, month):
    """해당 귀속월에 포함될 직원 목록 (예외자 제외)"""
    total_days = _cal.monthrange(year, month)[1]
    month_start = f'{year:04d}-{month:02d}-01'
    month_end   = f'{year:04d}-{month:02d}-{total_days:02d}'
    result = []
    for emp in db.get_all_employees():
        if emp['is_exception'] == 'Y':
            continue
        resign_d = (emp['resign_date'] or '')[:10]
        join_d   = (emp['join_date']   or '')[:10]
        if resign_d and resign_d < month_start:
            continue   # 이미 전월 이전 퇴사
        if join_d and join_d > month_end:
            continue   # 아직 입사 전
        result.append(emp)
    return result


def _build_payroll_entries(period_id, year, month, employees):
    """직원 목록으로 급여대장 기본 항목 생성"""
    total_days = _cal.monthrange(year, month)[1]
    ym = f'{year:04d}-{month:02d}'
    for emp in employees:
        monthly  = round((emp['annual_salary'] or 0) / 12)
        base     = emp['base_salary'] or 0
        overtime = max(0, monthly - base)

        join_d   = (emp['join_date']   or '')[:10]
        resign_d = (emp['resign_date'] or '')[:10]
        is_new   = 1 if join_d[:7]   == ym else 0
        is_res   = 1 if resign_d[:7] == ym else 0

        if is_new:
            work_days = total_days - int(join_d[8:10]) + 1
        elif is_res:
            work_days = int(resign_d[8:10])
        else:
            work_days = total_days

        if is_new or is_res:
            ratio    = work_days / total_days
            base     = round(base * ratio)
            overtime = round(overtime * ratio)

        gross = base + overtime
        db.create_payroll_entry({
            'period_id':     period_id,
            'employee_id':   emp['id'],
            'scheduled_days': total_days,
            'work_days':     work_days,
            'is_new_hire':   is_new,
            'is_resigned':   is_res,
            'base_salary':   base,
            'overtime_pay':  overtime,
            'gross_pay':     gross,
            'net_pay':       gross,
        })


@app.route('/payroll')
def payroll():
    periods = db.get_all_payroll_periods()
    return render_template('payroll/index.html', periods=periods)


@app.route('/payroll/new', methods=['GET', 'POST'])
def payroll_new():
    today = _date.today()
    if request.method == 'POST':
        year  = int(request.form['year'])
        month = int(request.form['month'])
        payment_date = request.form.get('payment_date', '')
        mode  = request.form.get('mode', 'new')

        if db.get_payroll_period_by_ym(year, month):
            flash(f'{year}년 {month}월 귀속 급여대장이 이미 존재합니다.', 'warning')
            return redirect(url_for('payroll'))

        period_id = db.create_payroll_period(year, month, payment_date)

        if mode == 'copy':
            prev_y = year if month > 1 else year - 1
            prev_m = month - 1 if month > 1 else 12
            prev   = db.get_payroll_period_by_ym(prev_y, prev_m)
            if prev:
                prev_entries = db.get_payroll_entries(prev['id'])
                emp_ids = [e['employee_id'] for e in prev_entries]
                employees = [db.get_employee(eid) for eid in emp_ids]
                employees = [e for e in employees if e and
                             e['is_exception'] != 'Y' and
                             ((e['resign_date'] or '') == '' or
                              e['resign_date'][:7] >= f'{year:04d}-{month:02d}')]
            else:
                flash('이전 귀속기간이 없어 신규로 생성합니다.', 'info')
                employees = _get_active_employees_for_period(year, month)
        else:
            employees = _get_active_employees_for_period(year, month)

        _build_payroll_entries(period_id, year, month, employees)
        db.update_payroll_period_status(period_id, 'step1')
        flash(f'{year}년 {month}월 급여대장 생성 완료 (총 {len(employees)}명)', 'success')
        return redirect(url_for('payroll_detail', period_id=period_id))

    prev_y = today.year if today.month > 1 else today.year - 1
    prev_m = today.month - 1 if today.month > 1 else 12
    has_prev = bool(db.get_payroll_period_by_ym(prev_y, prev_m))
    return render_template('payroll/new.html', today=today, has_prev=has_prev)


@app.route('/payroll/<int:period_id>')
def payroll_detail(period_id):
    period = db.get_payroll_period(period_id)
    if not period:
        flash('급여대장을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('payroll'))
    entries    = db.get_payroll_entries(period_id)
    allowances = db.get_all_allowances()
    paid_map   = db.get_payroll_allowances_by_period(period_id)
    ded_rows   = db.get_payroll_deductions_by_period(period_id)
    ded_names  = list(dict.fromkeys([d['deduction_name'] for d in ded_rows]))
    ded_map    = {}
    for d in ded_rows:
        ded_map.setdefault(d['entry_id'], {})[d['deduction_name']] = d['amount']
    return render_template('payroll/detail.html',
                           period=period, entries=entries,
                           allowances=allowances, paid_map=paid_map,
                           ded_names=ded_names, ded_map=ded_map)


@app.route('/payroll/<int:period_id>/step2', methods=['GET', 'POST'])
def payroll_step2(period_id):
    period = db.get_payroll_period(period_id)
    if not period:
        return redirect(url_for('payroll'))
    entries     = db.get_payroll_entries(period_id)
    total_days  = _cal.monthrange(period['year'], period['month'])[1]
    targets     = [e for e in entries if e['is_new_hire'] or e['is_resigned']]

    if request.method == 'POST':
        # 건너뛰기 (입퇴사자 없음)
        if request.form.get('_skip'):
            db.update_payroll_period_status(period_id, 'step2')
            return redirect(url_for('payroll_step3', period_id=period_id))

        for e in targets:
            wd = int(request.form.get(f'work_days_{e["id"]}', e['work_days']))
            wd = max(1, min(wd, total_days))
            emp  = db.get_employee(e['employee_id'])
            monthly  = round((emp['annual_salary'] or 0) / 12)
            base_full = emp['base_salary'] or 0
            ot_full   = max(0, monthly - base_full)
            ratio = wd / total_days
            base_p = round(base_full * ratio)
            ot_p   = round(ot_full   * ratio)
            db.update_payroll_entry(e['id'], {
                'work_days':   wd,
                'is_new_hire': e['is_new_hire'],
                'is_resigned': e['is_resigned'],
                'base_salary': base_p,
                'overtime_pay': ot_p,
                'gross_pay':   base_p + ot_p,
                'notes':       request.form.get(f'notes_{e["id"]}', ''),
            })
        db.update_payroll_period_status(period_id, 'step2')
        flash('입퇴사자 일할계산이 반영되었습니다.', 'success')
        return redirect(url_for('payroll_detail', period_id=period_id))

    return render_template('payroll/step2.html',
                           period=period, targets=targets, total_days=total_days)


@app.route('/payroll/<int:period_id>/step3')
def payroll_step3(period_id):
    """수당 처리 진입 → 첫 번째 수당으로 리디렉트"""
    period = db.get_payroll_period(period_id)
    if not period:
        return redirect(url_for('payroll'))
    allowances = [a for a in db.get_all_allowances() if a['is_active']]
    if not allowances:
        # 활성 수당 없으면 step3 완료 처리 후 step4로
        db.update_payroll_period_status(period_id, 'step3')
        flash('활성 수당 항목이 없어 수당 처리를 건너뜁니다.', 'info')
        return redirect(url_for('payroll_step4', period_id=period_id))
    return redirect(url_for('payroll_step3_item',
                            period_id=period_id,
                            allowance_id=allowances[0]['id']))


@app.route('/payroll/<int:period_id>/step3/<int:allowance_id>', methods=['GET', 'POST'])
def payroll_step3_item(period_id, allowance_id):
    """수당 항목 1개씩 처리"""
    period = db.get_payroll_period(period_id)
    if not period:
        return redirect(url_for('payroll'))

    entries    = db.get_payroll_entries(period_id)
    allowances = [a for a in db.get_all_allowances() if a['is_active']]
    allowance  = db.get_allowance(allowance_id)
    total_days = _cal.monthrange(period['year'], period['month'])[1]

    if not allowance or not allowance['is_active']:
        return redirect(url_for('payroll_step3', period_id=period_id))

    # 순서 계산
    allow_ids   = [a['id'] for a in allowances]
    cur_idx     = allow_ids.index(allowance_id) if allowance_id in allow_ids else 0
    prev_allow  = allowances[cur_idx - 1] if cur_idx > 0 else None
    next_allow  = allowances[cur_idx + 1] if cur_idx < len(allowances) - 1 else None
    is_last     = (next_allow is None)

    # 전월 데이터
    prev_y = period['year'] if period['month'] > 1 else period['year'] - 1
    prev_m = period['month'] - 1 if period['month'] > 1 else 12
    prev   = db.get_payroll_period_by_ym(prev_y, prev_m)
    prev_paid_map = db.get_payroll_allowances_by_period(prev['id']) if prev else {}

    if request.method == 'POST':
        for entry in entries:
            amt_str = request.form.get(f'amt_{entry["id"]}', '0').replace(',', '')
            amount  = int(amt_str or 0)
            is_paid = 1 if request.form.get(f'paid_{entry["id"]}') else 0
            notes   = request.form.get(f'notes_{entry["id"]}', '') or None
            db.upsert_payroll_allowance(entry['id'], allowance_id, amount, is_paid, notes)
        for entry in entries:
            db.recalculate_entry_totals(entry['id'])

        if is_last:
            db.update_payroll_period_status(period_id, 'step3')
            flash('수당 처리가 모두 완료되었습니다. 공제 처리로 이동하세요.', 'success')
            return redirect(url_for('payroll_detail', period_id=period_id))
        else:
            flash(f'[{allowance["name"]}] 저장 완료.', 'success')
            return redirect(url_for('payroll_step3_item',
                                    period_id=period_id,
                                    allowance_id=next_allow['id']))

    # GET — enrolled/available 분리 + 값 준비
    current_paid_map = db.get_payroll_allowances_by_period(period_id)
    load_prev = request.args.get('load_prev') == '1'

    emp_ids     = [e['employee_id'] for e in entries]
    ea_for_item = db.get_employee_allowances_for_item(emp_ids, allowance_id)

    enrolled_entries  = []
    available_entries = []
    for entry in entries:
        base_amt = ea_for_item.get(entry['employee_id'], 0)
        saved    = current_paid_map.get(entry['id'], {}).get(allowance_id)
        is_enrolled = (base_amt > 0) or bool(saved and saved.get('is_paid', 0) == 1)
        if is_enrolled:
            enrolled_entries.append(entry)
        else:
            available_entries.append(entry)

    amounts = {}
    for entry in enrolled_entries:
        emp_ea = {ea['allowance_item_id']: ea['amount']
                  for ea in db.get_employee_allowances(entry['employee_id'])}

        if load_prev and prev and prev_paid_map:
            prev_data = _find_prev_entry_allowance(
                prev_paid_map, prev['id'], entry['employee_id'], allowance_id)
            val = dict(prev_data) if prev_data else {
                'amount': emp_ea.get(allowance_id, 0), 'is_paid': 1, 'notes': ''}
        elif entry['id'] in current_paid_map and allowance_id in current_paid_map[entry['id']]:
            val = dict(current_paid_map[entry['id']][allowance_id])
            val.setdefault('notes', '')
        else:
            val = {'amount': emp_ea.get(allowance_id, 0), 'is_paid': 1, 'notes': ''}

        # 15일 조건 자동 플래그
        flag = False
        if (entry['is_new_hire'] or entry['is_resigned']) and \
           allowance['payment_condition'] == 'min_days' and \
           entry['work_days'] < allowance['condition_value']:
            val['is_paid'] = 0
            flag = True
        val['flag'] = flag
        amounts[entry['id']] = val

    return render_template('payroll/step3_item.html',
                           period=period,
                           entries=enrolled_entries,
                           available_entries=available_entries,
                           allowance=allowance, allowances=allowances,
                           cur_idx=cur_idx, amounts=amounts,
                           total_days=total_days,
                           prev_allow=prev_allow, next_allow=next_allow,
                           is_last=is_last,
                           has_prev=bool(prev),
                           load_prev=load_prev)


@app.route('/payroll/<int:period_id>/step3/<int:allowance_id>/add', methods=['POST'])
def payroll_step3_add_employee(period_id, allowance_id):
    """수당에 직원 추가 → 기본정보(employee_allowances)에도 반영"""
    entry_id = int(request.form.get('entry_id', 0))
    amount   = int(request.form.get('amount', '0').replace(',', '') or 0)
    entry    = db.get_payroll_entry(entry_id)
    if entry and amount > 0:
        db.upsert_payroll_allowance(entry_id, allowance_id, amount, 1)
        db.recalculate_entry_totals(entry_id)
        db.upsert_employee_allowance(entry['employee_id'], allowance_id, amount)
        flash('추가되었습니다. 기본정보에도 반영되어 다음달부터 고정 지급됩니다.', 'success')
    else:
        flash('금액을 0보다 크게 입력해주세요.', 'warning')
    return redirect(url_for('payroll_step3_item',
                            period_id=period_id, allowance_id=allowance_id))


@app.route('/payroll/<int:period_id>/step3/<int:allowance_id>/remove/<int:entry_id>', methods=['POST'])
def payroll_step3_remove_employee(period_id, allowance_id, entry_id):
    """수당에서 직원 영구 제외 → 기본정보에서도 제외"""
    entry = db.get_payroll_entry(entry_id)
    if entry:
        db.upsert_payroll_allowance(entry_id, allowance_id, 0, 0)
        db.recalculate_entry_totals(entry_id)
        db.upsert_employee_allowance(entry['employee_id'], allowance_id, 0)
        flash('제외되었습니다. 기본정보에서도 제외되어 다음달부터 미지급 처리됩니다.', 'success')
    return redirect(url_for('payroll_step3_item',
                            period_id=period_id, allowance_id=allowance_id))


def _refresh_entry_from_employee(entry):
    """직원 DB 기준으로 급여대장 기본급/연장수당 재적용 (수당·공제 건드리지 않음)"""
    emp = db.get_employee(entry['employee_id'])
    if not emp:
        return
    total_days = entry['scheduled_days'] or 1
    work_days  = entry['work_days']

    monthly    = round((emp['annual_salary'] or 0) / 12)
    base_full  = emp['base_salary'] or 0
    ot_full    = max(0, monthly - base_full)

    if entry['is_new_hire'] or entry['is_resigned']:
        ratio = work_days / total_days
        base  = round(base_full * ratio)
        ot    = round(ot_full   * ratio)
    else:
        base = base_full
        ot   = ot_full

    import sqlite3 as _s3
    conn = db.get_db()
    conn.execute(
        "UPDATE payroll_entries SET base_salary=?, overtime_pay=? WHERE id=?",
        (base, ot, entry['id'])
    )
    conn.commit()
    conn.close()
    db.recalculate_entry_totals(entry['id'])


@app.route('/payroll/<int:period_id>/step4/<int:deduction_id>/upload', methods=['POST'])
def payroll_step4_upload_insurance(period_id, deduction_id):
    """4대보험 고지서 엑셀 업로드 → match_key로 매칭 후 공제금액 자동 입력"""
    period    = db.get_payroll_period(period_id)
    deduction = db.get_deduction(deduction_id)
    if not period or not deduction:
        return redirect(url_for('payroll'))

    file = request.files.get('file')
    if not file or not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xls/.xlsx)을 업로드해주세요.', 'danger')
        return redirect(url_for('payroll_step4_item',
                                period_id=period_id, deduction_id=deduction_id))

    ded_name = deduction['name']
    is_health = any(k in ded_name for k in ['건강', '장기요양', '요양'])

    try:
        if is_health:
            _upload_health_insurance(period_id, file)
        else:
            _upload_pension_insurance(period_id, deduction, file)
    except Exception as e:
        flash(f'파일 처리 오류: {str(e)}', 'danger')

    return redirect(url_for('payroll_step4_item',
                            period_id=period_id, deduction_id=deduction_id))


def _read_insurance_rows(file):
    """업로드된 파일을 행 리스트로 읽기 (.xls/.xlsx 모두 지원)"""
    fname = file.filename.lower()
    if fname.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file, data_only=True)
        return list(wb.active.iter_rows(min_row=1, values_only=True))
    else:
        try:
            import xlrd as _xlrd
        except ImportError:
            raise RuntimeError('xlrd 라이브러리가 필요합니다. pip install xlrd 후 재시도하세요.')
        wb = _xlrd.open_workbook(file_contents=file.read())
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(i, j) for j in range(ws.ncols)] for i in range(ws.nrows)]


def _build_match_map(period_id):
    """period의 match_key → entry_id 맵 + 중복키 집합 반환"""
    from collections import Counter
    entries_info = db.get_payroll_entries_with_match_key(period_id)
    key_to_entry = {}
    counts = Counter()
    for ei in entries_info:
        mk = (ei['match_key'] or '').strip()
        if mk:
            key_to_entry[mk] = ei['entry_id']
            counts[mk] += 1
    duplicate_keys = {k for k, c in counts.items() if c > 1}
    return key_to_entry, duplicate_keys


def _upload_pension_insurance(period_id, deduction, file):
    """국민연금 고지서 업로드 (xlsx, col C=주민번호 마스킹, col F=직접납부기여금)"""
    rows = _read_insurance_rows(file)
    key_to_entry, duplicate_keys = _build_match_map(period_id)

    matched, unmatched, manual_needed = 0, [], []

    for row in rows:
        if len(row) < 6:
            continue
        raw_id = str(row[2]).strip() if row[2] else ''
        if '-' not in raw_id or len(raw_id) < 8:
            continue
        match_key = raw_id.replace('-', '')[:7]
        if not match_key.isdigit() or len(match_key) < 7:
            continue

        name = str(row[1]).strip() if row[1] else '이름미상'

        if match_key in duplicate_keys:
            amt_raw = row[5]
            amt = int(amt_raw) if isinstance(amt_raw, (int, float)) and amt_raw > 0 else 0
            manual_needed.append(f'{name}({match_key}): {amt:,}원')
            continue

        amt_raw = row[5]
        if not isinstance(amt_raw, (int, float)) or amt_raw <= 0:
            continue
        amt = int(amt_raw)

        if match_key in key_to_entry:
            entry_id = key_to_entry[match_key]
            db.upsert_payroll_deduction(entry_id, deduction['name'], amt)
            db.recalculate_entry_totals(entry_id)
            matched += 1
        else:
            unmatched.append(f'{name}({match_key}): {amt:,}원')

    _flash_upload_result(deduction['name'], matched, unmatched, manual_needed)


def _upload_health_insurance(period_id, file):
    """건강보험 고지서 업로드 (xls, col H=주민번호 풀, col R=건강고지, col AA=요양고지)
    건강보험 + 장기요양보험 공제 항목을 동시에 입력
    """
    # 건강보험 / 장기요양 공제 항목 탐색
    all_deds = db.get_all_deductions()
    health_ded = next((d for d in all_deds
                       if d['is_active'] and '건강' in d['name']
                       and '장기' not in d['name'] and '요양' not in d['name']), None)
    care_ded   = next((d for d in all_deds
                       if d['is_active'] and ('장기요양' in d['name'] or '요양' in d['name'])
                       and '건강' not in d['name']), None)

    if not health_ded:
        flash('건강보험 공제 항목을 찾을 수 없습니다. (항목명에 "건강" 포함 필요)', 'danger')
        return
    if not care_ded:
        flash('장기요양보험 공제 항목을 찾을 수 없습니다. (항목명에 "요양" 포함 필요)', 'danger')
        return

    rows = _read_insurance_rows(file)
    key_to_entry, duplicate_keys = _build_match_map(period_id)

    matched, unmatched, manual_needed = 0, [], []

    for row in rows:
        if len(row) < 27:
            continue
        # col H (index 7): 주민등록번호 (풀 13자리: XXXXXX-XXXXXXX)
        raw_id = str(row[7]).strip() if row[7] else ''
        if '-' not in raw_id or len(raw_id) < 13:
            continue
        match_key = raw_id.replace('-', '')[:7]
        if not match_key.isdigit() or len(match_key) < 7:
            continue

        name = str(row[6]).strip() if row[6] else '이름미상'

        # col R (index 17): 건강보험 고지금액 / col AA (index 26): 요양 고지보험료
        h_amt = int(row[17]) if isinstance(row[17], (int, float)) and row[17] > 0 else 0
        c_amt = int(row[26]) if isinstance(row[26], (int, float)) and row[26] > 0 else 0

        if h_amt == 0 and c_amt == 0:
            continue

        if match_key in duplicate_keys:
            manual_needed.append(
                f'{name}({match_key}): 건강 {h_amt:,}원 / 요양 {c_amt:,}원')
            continue

        if match_key in key_to_entry:
            entry_id = key_to_entry[match_key]
            if h_amt > 0:
                db.upsert_payroll_deduction(entry_id, health_ded['name'], h_amt)
            if c_amt > 0:
                db.upsert_payroll_deduction(entry_id, care_ded['name'], c_amt)
            db.recalculate_entry_totals(entry_id)
            matched += 1
        else:
            unmatched.append(
                f'{name}({match_key}): 건강 {h_amt:,}원 / 요양 {c_amt:,}원')

    label = f'{health_ded["name"]} + {care_ded["name"]}'
    _flash_upload_result(label, matched, unmatched, manual_needed)


def _flash_upload_result(label, matched, unmatched, manual_needed):
    """업로드 결과 flash 메시지 생성"""
    parts = [f'✅ {matched}명 [{label}] 자동 입력 완료']
    if manual_needed:
        parts.append(f'⚠️ 주민번호 중복 — 수동 입력 필요 {len(manual_needed)}명: '
                     + ' / '.join(manual_needed[:3]))
    if unmatched:
        parts.append(f'💰 미매칭(퇴사자 등) {len(unmatched)}명: '
                     + ' / '.join(unmatched[:5]))
    category = 'warning' if (manual_needed or unmatched) else 'success'
    flash('  |  '.join(parts), category)


@app.route('/payroll/<int:period_id>/refresh', methods=['POST'])
def payroll_refresh(period_id):
    """직원 DB에서 급여 기본값 재적용 (기본급·연장수당만, 수당·공제 보존)"""
    period = db.get_payroll_period(period_id)
    if not period:
        return redirect(url_for('payroll'))
    entries = db.get_payroll_entries(period_id)
    for entry in entries:
        _refresh_entry_from_employee(entry)
    flash(f'{len(entries)}명 기본급·연장수당을 직원 DB 기준으로 재적용했습니다.', 'success')
    next_url = request.form.get('next') or url_for('payroll_detail', period_id=period_id)
    return redirect(next_url)


def _find_prev_entry_allowance(prev_paid_map, prev_period_id, employee_id, allowance_id):
    """전월 급여대장에서 같은 직원의 수당 금액 찾기"""
    if not prev_period_id:
        return None
    conn = db.get_db()
    row = conn.execute(
        "SELECT id FROM payroll_entries WHERE period_id=? AND employee_id=?",
        (prev_period_id, employee_id)
    ).fetchone()
    conn.close()
    if not row:
        return None
    entry_id = row['id']
    return prev_paid_map.get(entry_id, {}).get(allowance_id)


@app.route('/payroll/<int:period_id>/step4')
def payroll_step4(period_id):
    """공제 처리 진입 → 첫 번째 공제 항목으로 리디렉트"""
    period = db.get_payroll_period(period_id)
    if not period:
        return redirect(url_for('payroll'))
    ded_items = [d for d in db.get_all_deductions() if d['is_active']]
    if not ded_items:
        db.update_payroll_period_status(period_id, 'completed')
        flash('활성 공제 항목이 없어 바로 완료 처리됩니다.', 'info')
        return redirect(url_for('payroll_detail', period_id=period_id))
    return redirect(url_for('payroll_step4_item',
                            period_id=period_id,
                            deduction_id=ded_items[0]['id']))


@app.route('/payroll/<int:period_id>/step4/<int:deduction_id>', methods=['GET', 'POST'])
def payroll_step4_item(period_id, deduction_id):
    """공제 항목 1개씩 처리"""
    period    = db.get_payroll_period(period_id)
    if not period:
        return redirect(url_for('payroll'))

    entries   = db.get_payroll_entries(period_id)
    ded_items = [d for d in db.get_all_deductions() if d['is_active']]
    deduction = db.get_deduction(deduction_id)

    if not deduction or not deduction['is_active']:
        return redirect(url_for('payroll_step4', period_id=period_id))

    ded_ids  = [d['id'] for d in ded_items]
    cur_idx  = ded_ids.index(deduction_id) if deduction_id in ded_ids else 0
    prev_ded = ded_items[cur_idx - 1] if cur_idx > 0 else None
    next_ded = ded_items[cur_idx + 1] if cur_idx < len(ded_items) - 1 else None
    is_last  = (next_ded is None)

    if request.method == 'POST':
        for entry in entries:
            amt = int(request.form.get(f'ded_{entry["id"]}', '0').replace(',', '') or 0)
            if amt > 0:
                db.upsert_payroll_deduction(entry['id'], deduction['name'], amt)
            else:
                db.delete_payroll_deduction_by_name(entry['id'], deduction['name'])
            db.recalculate_entry_totals(entry['id'])

        if is_last:
            db.update_payroll_period_status(period_id, 'completed')
            flash('공제 처리 완료. 급여대장이 확정되었습니다.', 'success')
            return redirect(url_for('payroll_detail', period_id=period_id))
        else:
            flash(f'[{deduction["name"]}] 저장 완료.', 'success')
            return redirect(url_for('payroll_step4_item',
                                    period_id=period_id,
                                    deduction_id=next_ded['id']))

    # GET — 직원별 공제금액 준비 (저장값 > 직원 기본값)
    saved = db.get_payroll_deductions_by_period(period_id)
    saved_by_entry = {}
    for s in saved:
        saved_by_entry.setdefault(s['entry_id'], {})[s['deduction_name']] = s['amount']

    ded_map = {}
    for entry in entries:
        emp_deds = {ed['deduction_item_id']: ed['amount']
                    for ed in db.get_employee_deductions(entry['employee_id'])}
        saved_val = saved_by_entry.get(entry['id'], {}).get(deduction['name'])
        ded_map[entry['id']] = saved_val if saved_val is not None else emp_deds.get(deduction_id, 0)

    # 고용보험: 저장값 없으면 총지급 × 0.8% 자동 계산
    if '고용' in deduction['name']:
        for entry in entries:
            if not ded_map[entry['id']]:
                ded_map[entry['id']] = round((entry['gross_pay'] or 0) * 0.008)

    return render_template('payroll/step4_item.html',
                           period=period, entries=entries,
                           deduction=deduction, ded_items=ded_items,
                           cur_idx=cur_idx, ded_map=ded_map,
                           prev_ded=prev_ded, next_ded=next_ded,
                           is_last=is_last)


@app.route('/payroll/<int:period_id>/export')
def payroll_export(period_id):
    """급여대장 엑셀 출력"""
    period = db.get_payroll_period(period_id)
    if not period:
        flash('급여대장을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('payroll'))

    entries    = db.get_payroll_entries(period_id)
    allowances = db.get_all_allowances()   # 전체 (비활성 포함 — 저장된 데이터 매핑용)
    paid_map   = db.get_payroll_allowances_by_period(period_id)   # {entry_id: {aid: {amount, is_paid}}}
    ded_rows   = db.get_payroll_deductions_by_period(period_id)   # list of {entry_id, deduction_name, amount}

    # 공제 컬럼 목록 (저장된 데이터 기준으로 정렬 보장)
    ded_items_ordered = db.get_all_deductions()
    ded_names = [d['name'] for d in ded_items_ordered if d['is_active']]
    # 혹시 저장된 공제명 중 목록에 없는 것도 추가
    saved_ded_names = list(dict.fromkeys([r['deduction_name'] for r in ded_rows]))
    for n in saved_ded_names:
        if n not in ded_names:
            ded_names.append(n)

    # 공제 맵: {entry_id: {deduction_name: amount}}
    ded_map = {}
    for r in ded_rows:
        ded_map.setdefault(r['entry_id'], {})[r['deduction_name']] = r['amount']

    # 수당 컬럼 목록 (활성화된 것 + 저장된 데이터에 있는 것)
    active_allowance_ids = {a['id']: a['name'] for a in allowances if a['is_active']}

    company_name = db.get_setting('company_name') or '회사명'

    # ── 워크북 생성 ──────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{period["year"]}년{period["month"]:02d}월_급여대장'

    # 스타일 정의
    thin  = Side(style='thin',   color='CCCCCC')
    thick = Side(style='medium', color='888888')
    def cell_border(thick_left=False, thick_right=False):
        l = thick if thick_left  else thin
        r = thick if thick_right else thin
        return Border(left=l, right=r, top=thin, bottom=thin)

    purple_fill  = PatternFill('solid', fgColor='7C3AED')
    violet_fill  = PatternFill('solid', fgColor='5B21B6')
    indigo_fill  = PatternFill('solid', fgColor='4C1D95')
    green_fill   = PatternFill('solid', fgColor='059669')
    gray_fill    = PatternFill('solid', fgColor='F3F4F6')
    hire_fill    = PatternFill('solid', fgColor='F0FFF4')
    resign_fill  = PatternFill('solid', fgColor='FFF5F5')
    white_fill   = PatternFill('solid', fgColor='FFFFFF')

    bold_white  = Font(bold=True, color='FFFFFF', size=10)
    bold_black  = Font(bold=True, color='1A202C', size=10)
    normal_font = Font(size=10)
    code_font   = Font(name='Courier New', size=9, color='1A4F8A')
    total_font  = Font(bold=True, size=10, color='1A202C')

    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')
    left_a = Alignment(horizontal='left',   vertical='center')

    # ── 행 1: 회사명 + 귀속기간 ─────────────────────────────
    total_cols = 7 + len(active_allowance_ids) + 2 + len(ded_names) + 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f'{company_name}  {period["year"]}년 {period["month"]}월 급여대장')
    title_cell.font      = Font(bold=True, size=14, color='1A202C')
    title_cell.alignment = left_a
    ws.row_dimensions[1].height = 28

    # ── 행 2: 지급일 + 인원 ─────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    payment_str = period['payment_date'] or '—'
    ws.cell(row=2, column=1,
            value=f'지급일: {payment_str}    총 {len(entries)}명    (단위: 원)').font = Font(size=10, color='718096')
    ws.row_dimensions[2].height = 18

    # ── 행 3: 헤더 ─────────────────────────────────────────
    FIXED_COLS   = ['이름코드', '소속', '구분', '소정일수', '근무일수', '기본급', '연장수당']
    allow_names  = [active_allowance_ids[aid] for aid in active_allowance_ids]
    ALLOW_COLS   = allow_names + ['수당합계', '총지급']
    DED_COLS     = ded_names   + ['공제합계', '실지급']
    ALL_HEADERS  = FIXED_COLS + ALLOW_COLS + DED_COLS

    header_row = 3
    ws.row_dimensions[header_row].height = 22
    col_idx = 1
    # 고정 컬럼
    for h in FIXED_COLS:
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.fill = c.fill = purple_fill
        c.font = bold_white
        c.alignment = center
        c.border = cell_border()
        col_idx += 1
    # 수당 컬럼
    for h in allow_names:
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.fill = purple_fill
        c.font = bold_white
        c.alignment = center
        c.border = cell_border()
        col_idx += 1
    # 수당합계
    c = ws.cell(row=header_row, column=col_idx, value='수당합계')
    c.fill = violet_fill; c.font = bold_white; c.alignment = center; c.border = cell_border(thick_left=True)
    col_idx += 1
    # 총지급
    c = ws.cell(row=header_row, column=col_idx, value='총지급')
    c.fill = violet_fill; c.font = bold_white; c.alignment = center; c.border = cell_border()
    col_total_gross = col_idx
    col_idx += 1
    # 공제 컬럼
    for h in ded_names:
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.fill = purple_fill; c.font = bold_white; c.alignment = center; c.border = cell_border(thick_left=(h == ded_names[0]) if ded_names else False)
        col_idx += 1
    # 공제합계
    c = ws.cell(row=header_row, column=col_idx, value='공제합계')
    c.fill = indigo_fill; c.font = bold_white; c.alignment = center; c.border = cell_border(thick_left=True)
    col_idx += 1
    # 실지급
    c = ws.cell(row=header_row, column=col_idx, value='실지급')
    c.fill = indigo_fill; c.font = bold_white; c.alignment = center; c.border = cell_border()

    # ── 데이터 행 ────────────────────────────────────────────
    # 합계 누적용
    sum_base = sum_ot = sum_allow_total = sum_gross = sum_net = 0
    sum_allow_by_id  = {aid: 0 for aid in active_allowance_ids}
    sum_ded_by_name  = {n: 0 for n in ded_names}
    sum_ded_total    = 0

    for data_row, entry in enumerate(entries, start=4):
        ws.row_dimensions[data_row].height = 18
        row_fill = hire_fill if entry['is_new_hire'] else (resign_fill if entry['is_resigned'] else white_fill)

        col_idx = 1

        def wc(val, align=right, fmt=None, bold=False, fill=None):
            nonlocal col_idx
            c = ws.cell(row=data_row, column=col_idx, value=val)
            c.alignment = align
            c.font = Font(bold=bold, size=10)
            c.fill = fill or row_fill
            c.border = cell_border()
            if fmt:
                c.number_format = fmt
            col_idx += 1
            return c

        # 이름코드
        c = ws.cell(row=data_row, column=col_idx, value=entry['name_code'])
        c.font = code_font; c.alignment = left_a; c.fill = row_fill; c.border = cell_border(); col_idx += 1
        # 소속
        wc(entry['department'] or '—', align=left_a)
        # 구분
        label = '입사' if entry['is_new_hire'] else ('퇴사' if entry['is_resigned'] else '재직')
        wc(label, align=center)
        # 소정일수 / 근무일수
        wc(entry['scheduled_days'], align=center)
        wc(entry['work_days'],      align=center)
        # 기본급 / 연장수당
        wc(entry['base_salary'],   fmt='#,##0')
        wc(entry['overtime_pay'],  fmt='#,##0')

        sum_base += entry['base_salary']
        sum_ot   += entry['overtime_pay']

        # 수당 컬럼
        allow_total_entry = 0
        for aid in active_allowance_ids:
            amt_data = paid_map.get(entry['id'], {}).get(aid)
            amt = amt_data['amount'] if (amt_data and amt_data['is_paid']) else 0
            wc(amt or None, fmt='#,##0')
            sum_allow_by_id[aid] += amt
            allow_total_entry    += amt
        # 수당합계
        c = ws.cell(row=data_row, column=col_idx, value=allow_total_entry or None)
        c.font = Font(bold=True, size=10); c.alignment = right; c.fill = row_fill
        c.border = cell_border(thick_left=True); c.number_format = '#,##0'; col_idx += 1
        sum_allow_total += allow_total_entry
        # 총지급
        gross = entry['gross_pay']
        c = ws.cell(row=data_row, column=col_idx, value=gross)
        c.font = Font(bold=True, size=10); c.alignment = right; c.fill = row_fill
        c.border = cell_border(); c.number_format = '#,##0'; col_idx += 1
        sum_gross += gross

        # 공제 컬럼
        ded_total_entry = 0
        first_ded = True
        for dn in ded_names:
            amt = ded_map.get(entry['id'], {}).get(dn, 0)
            c = ws.cell(row=data_row, column=col_idx, value=amt or None)
            c.alignment = right; c.fill = row_fill; c.number_format = '#,##0'
            c.border = cell_border(thick_left=first_ded); c.font = normal_font
            col_idx += 1; first_ded = False
            sum_ded_by_name[dn] += amt
            ded_total_entry     += amt
        # 공제합계
        c = ws.cell(row=data_row, column=col_idx, value=ded_total_entry or None)
        c.font = Font(bold=True, size=10); c.alignment = right; c.fill = row_fill
        c.border = cell_border(thick_left=True); c.number_format = '#,##0'; col_idx += 1
        sum_ded_total += ded_total_entry
        # 실지급
        net = entry['net_pay']
        c = ws.cell(row=data_row, column=col_idx, value=net)
        c.font = Font(bold=True, size=10, color='4C1D95'); c.alignment = right; c.fill = row_fill
        c.border = cell_border(); c.number_format = '#,##0'
        sum_net += net

    # ── 합계 행 ─────────────────────────────────────────────
    footer_row = 4 + len(entries)
    ws.row_dimensions[footer_row].height = 20
    col_idx = 1

    def fc(val, fmt=None, bold=True, thick_l=False, color='1A202C'):
        nonlocal col_idx
        c = ws.cell(row=footer_row, column=col_idx, value=val)
        c.font  = Font(bold=bold, size=10, color=color)
        c.fill  = gray_fill
        c.alignment = right
        c.border = cell_border(thick_left=thick_l)
        if fmt:
            c.number_format = fmt
        col_idx += 1
        return c

    # 합계 레이블 (이름코드~구분 병합)
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=3)
    lc = ws.cell(row=footer_row, column=1, value='합  계')
    lc.font = Font(bold=True, size=10); lc.alignment = center; lc.fill = gray_fill
    lc.border = cell_border()
    col_idx = 4

    # 소정일수/근무일수 — 공백
    fc(None); fc(None)
    # 기본급/연장
    fc(sum_base, fmt='#,##0')
    fc(sum_ot,   fmt='#,##0')
    # 수당 합계 행
    for aid in active_allowance_ids:
        fc(sum_allow_by_id[aid] or None, fmt='#,##0')
    fc(sum_allow_total or None, fmt='#,##0', thick_l=True)
    fc(sum_gross, fmt='#,##0')
    # 공제
    first_ded = True
    for dn in ded_names:
        fc(sum_ded_by_name[dn] or None, fmt='#,##0', thick_l=first_ded)
        first_ded = False
    fc(sum_ded_total or None, fmt='#,##0', thick_l=True)
    fc(sum_net, fmt='#,##0', color='4C1D95')

    # ── 열 너비 자동 조정 ─────────────────────────────────────
    col_widths = [16, 12, 8, 8, 8, 14, 12]  # 고정 컬럼
    for _ in active_allowance_ids:
        col_widths.append(13)
    col_widths += [11, 14]   # 수당합계, 총지급
    for _ in ded_names:
        col_widths.append(13)
    col_widths += [11, 14]   # 공제합계, 실지급

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 틀고정: 헤더 아래, 이름코드 우측
    ws.freeze_panes = 'C4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'{period["year"]}년{period["month"]:02d}월_급여대장.xlsx'
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/payroll/<int:period_id>/delete', methods=['POST'])
def payroll_period_delete(period_id):
    period = db.get_payroll_period(period_id)
    if period:
        db.delete_payroll_period(period_id)
        flash(f'{period["year"]}년 {period["month"]}월 급여대장이 삭제되었습니다.', 'success')
    return redirect(url_for('payroll'))


@app.route('/labor-cost')
def labor_cost():
    return render_template('labor_cost/index.html')


@app.route('/analysis')
def analysis():
    return render_template('analysis/index.html')


# ════════════════════════════════════════════════════════
#  설정
# ════════════════════════════════════════════════════════
@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        db.set_setting('company_name', request.form.get('company_name', ''))
        db.set_setting('abnormal_threshold', request.form.get('abnormal_threshold', '20'))
        flash('설정이 저장되었습니다.', 'success')
        return redirect(url_for('settings'))
    cfg = db.get_all_settings()
    return render_template('settings/index.html', cfg=cfg)


# ════════════════════════════════════════════════════════
if __name__ == '__main__':
    print("=" * 50)
    print("  인건비 관리 시스템 시작")
    print("  브라우저에서 http://localhost:5000 접속")
    print("=" * 50)
    app.run(debug=True, port=5000)
